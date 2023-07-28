

from django.http import JsonResponse
from proyectos.serializers.signup import UserSignUpSerializer
from django.views.decorators.csrf import csrf_exempt
from proyectos.serializers.user import UserModelSerializer
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from proyectos.models import Perfil, Rol, Ficha, Inscrito
from django.contrib.auth.models import User
from openpyxl import load_workbook
import json

# # Vista para subir un archivo y procesar los datos de usuarios en él
# @csrf_exempt
# def upload_file(request):
#     if request.method == 'POST':
#         # Obtener ficha_id, rol_id y el archivo del cuerpo de la solicitud POST
#         ficha_id = request.POST.get('ficha_id', None)
#         rol_id = request.POST.get('rol_id', None)
#         file = request.FILES.get('file', None)
#         print('rol', rol_id)

#         # Verificar que los datos requeridos están presentes en la solicitud
#         if not file or ficha_id is None or rol_id is None:
#             return JsonResponse({'error': 'Invalid request data'}, status=status.HTTP_400_BAD_REQUEST)

#         try:
#             # Intentar obtener instancias de Rol y Ficha utilizando los IDs proporcionados
#             rol_instance = Rol.objects.get(id=rol_id)
#             ficha_instance = Ficha.objects.get(id=ficha_id)
#         except Rol.DoesNotExist:
#             return JsonResponse({'error': 'Invalid rol_id'}, status=status.HTTP_400_BAD_REQUEST)
#         except Ficha.DoesNotExist:
#             return JsonResponse({'error': 'Invalid ficha_id'}, status=status.HTTP_400_BAD_REQUEST)

#         # Cargar el archivo Excel utilizando la biblioteca openpyxl
#         workbook = load_workbook(file)
#         sheet = workbook.active
#         users = []
#         skip_first_row = True  # Flag to skip the first row

#         # Recorrer las filas del archivo Excel y procesar los datos de los usuarios
#         for index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
#             print(type(row[1]))
#             if skip_first_row:
#                 skip_first_row = False
#                 continue  # Skip processing the first row

#             try:
#                 # Extraer los datos del usuario de la fila actual
#                 row_data = {
#                     'username': str(row[0]),
#                     'email': str(row[0]),
#                     'first_name': str(row[3]),
#                     'last_name': str(row[4]),
#                     'password': str(row[1]),
#                     'password_confirmation': str(row[1])  # Suponemos que se utiliza la misma contraseña para confirmación
#                 }


#                 # Imprimir los datos del usuario de la fila actual (opcional, solo para depuración)
#                 print(f"Row {index} Data:", row_data)

#                 # Verificar que no haya valores vacíos o nulos en los datos del usuario
#                 if any(value is None for value in row_data.values()):
#                     print(f"Skipping Row {row_data} - Incomplete data.")
#                     break

#                 # Validar los datos del usuario utilizando el serializador UserSignUpSerializer
#                 serializer = UserSignUpSerializer(data=row_data)
#                 serializer.is_valid(raise_exception=True)
#                 user = serializer.save()
#                 user_data = UserModelSerializer(user).data
#                 user_id = user_data['id']

#                 try:
#                     # Intentar obtener la instancia del usuario recién creado
#                     user_instance = User.objects.get(id=user_id)
#                 except User.DoesNotExist:
#                     return JsonResponse({'error': f'User with ID {user_id} not found'}, status=status.HTTP_400_BAD_REQUEST)

#                 # Crear un perfil relacionado con el usuario y la ficha
#                 per = Perfil()
#                 per.documento = row[1]
#                 per.rol = rol_instance
#                 per.usuario = user_instance
#                 per.tipo_documento = row[2]
#                 per.save()

#                 # Crear un registro de Inscrito relacionando el perfil con la ficha
#                 inscrito = Inscrito()
#                 inscrito.perfil = per
#                 inscrito.ficha = ficha_instance
#                 inscrito.save()

#                 # Imprimir los datos de los registros creados (opcional, solo para depuración)
#                 print(inscrito)
#                 print(per)

#                 # Agregar los datos del usuario al resultado final
#                 users.append(user_data)

#             except StopIteration:
#                 # Romper el bucle cuando no hay más filas en el archivo Excel
#                 break

#         workbook.close()

#         # Devolver una respuesta JSON con los datos de los usuarios registrados
#         return JsonResponse(users, safe=False)

#     # Devolver una respuesta de error si el método de solicitud no es POST
#     return JsonResponse({'error': 'Invalid request method'}, status=status.HTTP_405_METHOD_NOT_ALLOWED)
@csrf_exempt
def upload_file(request):
    if request.method == 'POST':
        # Obtener ficha_id, rol_id y el archivo del cuerpo de la solicitud POST
        ficha_id = request.POST.get('ficha_id', None)
        rol_id = request.POST.get('rol_id', None)
        file = request.FILES.get('file', None)
        print('rol', rol_id)

        # Verificar que los datos requeridos están presentes en la solicitud
        if not file or ficha_id is None or rol_id is None:
            return JsonResponse({'error': 'Invalid request data'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            # Intentar obtener instancias de Rol y Ficha utilizando los IDs proporcionados
            rol_instance = Rol.objects.get(id=rol_id)
            ficha_instance = Ficha.objects.get(id=ficha_id)
        except Rol.DoesNotExist:
            return JsonResponse({'error': 'Invalid rol_id'}, status=status.HTTP_400_BAD_REQUEST)
        except Ficha.DoesNotExist:
            return JsonResponse({'error': 'Invalid ficha_id'}, status=status.HTTP_400_BAD_REQUEST)

        # Cargar el archivo Excel utilizando la biblioteca openpyxl
        workbook = load_workbook(file)
        sheet = workbook.active
        users = []
        skip_first_row = True  # Flag to skip the first row


        # Recorrer las filas del archivo Excel y procesar los datos de los usuarios
        for index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if skip_first_row:
                skip_first_row = False
                continue  # Skip processing the first row
            if row[0] == None:
                break
            password = str(int(row[1]))
            try:
                # Extraer los datos del usuario de la fila actual
                # Supongamos que el valor numérico en row[1] es 10279734.0
                
                

                # Ahora el valor de 'password' en row_data será "10279734" (como una cadena de texto sin puntos)
                row_data = {
                    'username': str(row[0]),
                    'email': str(row[0]),
                    'first_name': str(row[3]),
                    'last_name': str(row[4]),
                    'password': password ,
                    'password_confirmation':password  # Suponemos que se utiliza la misma contraseña para confirmación
                }
                # Imprimir los datos del usuario de la fila actual (opcional, solo para depuración)
                print(f"Row {index} Data:", row_data)

                try:
                    # Validar los datos del usuario utilizando el serializador UserSignUpSerializer
                    serializer = UserSignUpSerializer(data=row_data)
                    serializer.is_valid(raise_exception=True)
                    user = serializer.save()

                    # Crear el modelo Perfil y asignar valores a sus atributos
                    per = Perfil()
                    per.documento = row[1]
                    per.rol = rol_instance
                    per.usuario = user
                    per.tipo_documento = row[2]

                    # Guardar el modelo Perfil en la base de datos
                    per.save()

                    # Crear el modelo Inscrito y asignar valores a sus atributos
                    inscrito = Inscrito()
                    inscrito.perfil = per
                    inscrito.ficha = ficha_instance

                    # Guardar el modelo Inscrito en la base de datos
                    inscrito.save()

                    # Resto del código para perfil e inscrito...

                except Exception as e:
                    # Capturar el error (imprimirlo o usar un logger para guardar el error)
                    print(f"Error procesando Fila {index}: {str(e)}")
                    continue  # Continuar con la siguiente iteración si ocurre un error al guardar Perfil o Inscrito
                user_data = UserModelSerializer(user).data


                # Agregar los datos del usuario al resultado final
                users.append(user_data)

            except StopIteration:
                # Romper el bucle cuando no hay más filas en el archivo Excel
                break

        workbook.close()

        # Devolver una respuesta JSON con los datos de los usuarios registrados
        return JsonResponse(users, safe=False)

    # Devolver una respuesta de error si el método de solicitud no es POST
    return JsonResponse({'error': 'Invalid request method'}, status=status.HTTP_405_METHOD_NOT_ALLOWED)